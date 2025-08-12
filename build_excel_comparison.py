#!/usr/bin/env python
"""Generate Excel workbook comparing ground truth vs extracted fields side by side.

Output: debug/benchmark_side_by_side.xlsx
Sheets:
  - KAEK: ΑΔΑ, GT ΚΑΕΚ, pdfplumber ΚΑΕΚ, docling ΚΑΕΚ, match flags
  - Owners: one row per owner slot (paired by index) with normalized comparison
  - Coverage: one row per (ΑΔΑ, Group, Key) showing GT value and both extractors

Assumptions:
  * Ground truth CSV path: data/01_benchmark/eadeies_final.csv
  * Structured JSON directory: debug/structured_json

Usage:
  ./.venv/bin/python build_excel_comparison.py \
      --benchmark-csv data/01_benchmark/eadeies_final.csv \
      --structured-dir debug/structured_json \
      --out debug/benchmark_side_by_side.xlsx
"""
from __future__ import annotations
import argparse
import csv
import json
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from benchmark_evaluation import normalize_kaek, equivalent_kaek, extract_ground_truth_owners, load_ground_truth, parse_float, GROUPS, COVERAGE_KEYS, normalize_owner_component


def _norm_header(s: str) -> str:
    """Canonicalize a ground-truth coverage header for matching.

    Replaces NBSP, collapses all whitespace runs to single spaces, strips.
    Also collapses multiple spaces around hyphens.
    """
    if s is None:
        return ""
    s = s.replace("\xa0", " ")
    # Collapse whitespace
    import re as _re
    s = _re.sub(r"\s+", " ", s)
    # Remove stray double spaces after hyphen patterns
    s = s.replace("-  ", "- ")
    return s.strip()


def load_structured(structured_dir: Path, stem: str, extractor: str) -> Dict[str, Any] | None:
    path = structured_dir / f"{stem}_{extractor}_structured.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def build_kaek_sheet(wb: Workbook, stems: List[str], gt_rows: Dict[str, dict], structured_dir: Path):
    ws = wb.create_sheet("KAEK")
    ws.append(["ΑΔΑ", "GT ΚΑΕΚ", "pdfplumber ΚΑΕΚ", "docling ΚΑΕΚ", "pdfplumber match", "docling match"])
    for stem in stems:
        gt = gt_rows.get(stem)
        if not gt:
            continue
        gt_kaek = normalize_kaek(gt.get("ΚΑΕΚ") or "")
        row = [stem, gt_kaek]
        for extractor in ["pdfplumber", "docling"]:
            data = load_structured(structured_dir, stem, extractor)
            pred = normalize_kaek(data.get("ΚΑΕΚ", "")) if data else ""
            row.append(pred)
        # matches
        pdf_match = equivalent_kaek(row[2], gt_kaek)
        doc_match = equivalent_kaek(row[3], gt_kaek)
        row.extend([int(pdf_match), int(doc_match)])
        ws.append(row)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)


def build_owners_sheet(wb: Workbook, stems: List[str], gt_rows: Dict[str, dict], structured_dir: Path):
    ws = wb.create_sheet("Owners")
    ws.append(["ΑΔΑ", "Index", "GT Surname", "GT Name", "pdfplumber Surname", "pdfplumber Name", "pdfplumber Match", "docling Surname", "docling Name", "docling Match"])    
    for stem in stems:
        gt = gt_rows.get(stem)
        if not gt:
            continue
        gt_owners = extract_ground_truth_owners(gt)
        max_len = max(len(gt_owners), 5)  # show at least 5 rows if predictions longer
        # Collect extractor owners (raw order)
        preds = {}
        for extractor in ["pdfplumber", "docling"]:
            data = load_structured(structured_dir, stem, extractor) or {}
            owners_data = data.get("Στοιχεία κυρίου του έργου", [])
            preds[extractor] = owners_data
            max_len = max(max_len, len(owners_data))
        for idx in range(max_len):
            gt_surn = gt_owners[idx].surname if idx < len(gt_owners) else ""
            gt_name = gt_owners[idx].name if idx < len(gt_owners) else ""
            pdf_surn = preds["pdfplumber"][idx].get("Επώνυμο/ία", "") if idx < len(preds["pdfplumber"]) else ""
            pdf_name = preds["pdfplumber"][idx].get("Όνομα", "") if idx < len(preds["pdfplumber"]) else ""
            doc_surn = preds["docling"][idx].get("Επώνυμο/ία", "") if idx < len(preds["docling"]) else ""
            doc_name = preds["docling"][idx].get("Όνομα", "") if idx < len(preds["docling"]) else ""
            # Compute matches (normalized) only if GT present
            if gt_surn or gt_name:
                gt_key = (normalize_owner_component(gt_surn), normalize_owner_component(gt_name))
                pdf_key = (normalize_owner_component(pdf_surn), normalize_owner_component(pdf_name)) if (pdf_surn or pdf_name) else ("","")
                doc_key = (normalize_owner_component(doc_surn), normalize_owner_component(doc_name)) if (doc_surn or doc_name) else ("","")
                pdf_match = int(pdf_key == gt_key) if (pdf_surn or pdf_name) else ""
                doc_match = int(doc_key == gt_key) if (doc_surn or doc_name) else ""
            else:
                pdf_match = doc_match = ""
            ws.append([stem, idx + 1, gt_surn, gt_name, pdf_surn, pdf_name, pdf_match, doc_surn, doc_name, doc_match])
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)


def build_coverage_sheet(wb: Workbook, stems: List[str], gt_rows: Dict[str, dict], structured_dir: Path):
    ws = wb.create_sheet("Coverage")
    ws.append(["ΑΔΑ", "Group", "Metric", "GT", "pdfplumber", "docling", "pdfplumber match", "docling match", "pdfplumber diff", "docling diff"])
    for stem in stems:
        gt = gt_rows.get(stem)
        if not gt:
            continue
        for group in GROUPS:
            for key in COVERAGE_KEYS:
                # Build canonical expected header and search using normalized comparison
                expected = _norm_header(f"{group} - {key}")
                gt_val_raw = None
                for h, v in gt.items():
                    if _norm_header(h) == expected:
                        gt_val_raw = v
                        break
                gt_val = parse_float(gt_val_raw) if gt_val_raw is not None else None
                if gt_val is None:
                    continue
                row = [stem, group, key, gt_val]
                pdf_val = doc_val = None
                for extractor in ["pdfplumber", "docling"]:
                    data = load_structured(structured_dir, stem, extractor)
                    val = None
                    if data:
                        val = data.get("Στοιχεία Διαγράμματος Κάλυψης", {}).get(group, {}).get(key)
                    if extractor == "pdfplumber":
                        pdf_val = val
                    else:
                        doc_val = val
                pdf_match = int(isinstance(pdf_val,(int,float)) and abs(pdf_val-gt_val)<1e-6) if pdf_val is not None else 0
                doc_match = int(isinstance(doc_val,(int,float)) and abs(doc_val-gt_val)<1e-6) if doc_val is not None else 0
                pdf_diff = (pdf_val - gt_val) if isinstance(pdf_val, (int, float)) and not pdf_match else 0.0 if pdf_match else None
                doc_diff = (doc_val - gt_val) if isinstance(doc_val, (int, float)) and not doc_match else 0.0 if doc_match else None
                row.extend([pdf_val, doc_val, pdf_match, doc_match, pdf_diff, doc_diff])
                ws.append(row)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)


def build_coverage_wide_sheet(wb: Workbook, stems: List[str], gt_rows: Dict[str, dict], structured_dir: Path):
    """Add a sheet with one row per stem and all coverage metrics as columns.

    Column naming: <Group abbrev>:<Metric> (abbrev first 3 Greek letters of group) for compactness.
    """
    ws = wb.create_sheet("CoverageWide")
    # Build header
    headers = ["ΑΔΑ"]
    col_keys = []  # store tuples (group, key)
    for group in GROUPS:
        for key in COVERAGE_KEYS:
            abbrev = group[:4]  # first few chars to distinguish
            h = f"{abbrev}:{key}"
            headers.append(h)
            col_keys.append((group, key))
    ws.append(headers)
    for stem in stems:
        gt = gt_rows.get(stem)
        if not gt:
            continue
        # Preload structured
        structured = {ext: load_structured(structured_dir, stem, ext) or {} for ext in ["pdfplumber", "docling"]}
        row = [stem]
        for group, key in col_keys:
            # Choose docling value first (better accuracy) fallback to pdfplumber if docling None
            doc_val = structured["docling"].get("Στοιχεία Διαγράμματος Κάλυψης", {}).get(group, {}).get(key)
            pdf_val = structured["pdfplumber"].get("Στοιχεία Διαγράμματος Κάλυψης", {}).get(group, {}).get(key)
            val = doc_val if isinstance(doc_val, (int, float)) else pdf_val
            row.append(val)
        ws.append(row)
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)


def build_coverage_mismatches_sheet(wb: Workbook, stems: List[str], gt_rows: Dict[str, dict], structured_dir: Path):
    """Sheet listing only coverage cells where prediction != ground truth for either extractor.

    Columns: ΑΔΑ, Extractor, Group, Metric, Ground Truth, Predicted, Diff
    """
    ws = wb.create_sheet("CoverageMismatches")
    ws.append(["ΑΔΑ", "Extractor", "Group", "Metric", "Ground Truth", "Pred", "Diff"])
    for stem in stems:
        gt = gt_rows.get(stem)
        if not gt:
            continue
        # Preload structured data per extractor
        structured = {ext: load_structured(structured_dir, stem, ext) or {} for ext in ["pdfplumber", "docling"]}
        for group in GROUPS:
            for key in COVERAGE_KEYS:
                expected = _norm_header(f"{group} - {key}")
                gt_val_raw = None
                for h, v in gt.items():
                    if _norm_header(h) == expected:
                        gt_val_raw = v
                        break
                gt_val = parse_float(gt_val_raw) if gt_val_raw is not None else None
                if gt_val is None:
                    continue
                for ext in ["pdfplumber", "docling"]:
                    pred = structured[ext].get("Στοιχεία Διαγράμματος Κάλυψης", {}).get(group, {}).get(key)
                    if isinstance(pred, (int, float)) and abs(pred - gt_val) < 1e-6:
                        continue
                    # Include mismatch (also include when pred is None)
                    diff = (pred - gt_val) if isinstance(pred, (int, float)) and gt_val is not None else None
                    ws.append([stem, ext, group, key, gt_val, pred, diff])
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)


def main():
    ap = argparse.ArgumentParser(description="Build Excel with side-by-side ground truth vs extracted values")
    ap.add_argument("--benchmark-csv", type=Path, default=Path("data/01_benchmark/eadeies_final.csv"))
    ap.add_argument("--structured-dir", type=Path, default=Path("debug/structured_json"))
    ap.add_argument("--out", type=Path, default=Path("debug/benchmark_side_by_side.xlsx"))
    args = ap.parse_args()

    gt_rows = load_ground_truth(args.benchmark_csv)
    stems = sorted(gt_rows.keys())

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    build_kaek_sheet(wb, stems, gt_rows, args.structured_dir)
    build_owners_sheet(wb, stems, gt_rows, args.structured_dir)
    build_coverage_sheet(wb, stems, gt_rows, args.structured_dir)
    build_coverage_wide_sheet(wb, stems, gt_rows, args.structured_dir)
    build_coverage_mismatches_sheet(wb, stems, gt_rows, args.structured_dir)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote Excel comparison to {args.out}")

if __name__ == "__main__":  # pragma: no cover
    main()
